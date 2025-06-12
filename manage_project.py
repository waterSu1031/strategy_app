import os
import ast
import pandas as pd
import chardet
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

# 🔧 다른 파일의 함수/클래스 이름 전체 수집 (Python 파일만 해당)
def extract_project_symbols(base_dir):
    symbols = set()
    for root, _, files in os.walk(base_dir):
        if any(ignored in root for ignored in ['.venv', '__pycache__', '.git']): # .git 추가
            continue
        for file in files:
            if file.endswith('.py'): # Python 파일만 AST 파싱하여 심볼 추출
                path = os.path.join(root, file)
                try:
                    with open(path, 'rb') as f:
                        raw = f.read()
                    encoding = chardet.detect(raw)['encoding'] or 'utf-8'
                    content = raw.decode(encoding, errors='ignore')
                    node = ast.parse(content)
                    for n in ast.walk(node):
                        if isinstance(n, (ast.FunctionDef, ast.ClassDef)):
                            symbols.add(n.name)
                except Exception:
                    continue
    return symbols

# 🆕 현재 파일에 정의된 함수/클래스 이름 수집 (Python 파일만 해당)
def extract_local_symbols(file_path):
    local = set()
    try:
        with open(file_path, 'rb') as f:
            raw = f.read()
        encoding = chardet.detect(raw)['encoding'] or 'utf-8'
        content = raw.decode(encoding, errors='ignore')
        node = ast.parse(content)
        for n in ast.walk(node):
            if isinstance(n, (ast.FunctionDef, ast.ClassDef)):
                local.add(n.name)
    except Exception:
        pass
    return local

# 📍 호출된 이름 추출 도우미
def get_call_name(n):
    if isinstance(n, ast.Name):
        return n.id
    elif isinstance(n, ast.Attribute):
        base = get_call_name(n.value)
        return f"{base}.{n.attr}" if base else n.attr
    return ""

# 💡 클래스/함수/Calls 정보 추출 (외부 파일 호출만 포함) (Python 파일 전용)
def extract_functions_and_classes(file_path, project_symbols, local_symbols):
    def read_file_content(path):
        with open(path, 'rb') as f:
            raw = f.read()
        encoding = chardet.detect(raw)['encoding'] or 'utf-8'
        try:
            return raw.decode(encoding, errors='ignore')
        except Exception as e:
            print(f"[Error] Failed to decode {path} using {encoding}: {e}")
            return ""

    file_content = read_file_content(file_path)
    if not file_content:
        return []

    try:
        node = ast.parse(file_content, filename=file_path)
    except SyntaxError as e:
        print(f"[SyntaxError] Skipping {file_path}: {e}")
        return []

    structure = []

    class FunctionVisitor(ast.NodeVisitor):
        def __init__(self, parent=''):
            self.parent = parent
            self.current_class = None

        def visit_ClassDef(self, class_node):
            structure.append({
                "Type": "Class",
                "Name": class_node.name,
                "Parent": self.parent,
                "Docstring": ast.get_docstring(class_node) or "",
                "Calls": ""
            })
            prev_class = self.current_class
            self.current_class = class_node.name
            for node in class_node.body:
                self.visit(node)
            self.current_class = prev_class

        def visit_FunctionDef(self, func_node):
            raw_calls = [get_call_name(n.func) for n in ast.walk(func_node) if isinstance(n, ast.Call)]
            filtered_calls = [
                c for c in raw_calls
                # project_symbols에 있는 호출 중, 현재 파일에 정의되지 않은 심볼만 포함
                if c.split('.')[0] in project_symbols and c.split('.')[0] not in local_symbols
            ]
            parent_name = self.current_class or self.parent
            structure.append({
                "Type": "Function",
                "Name": func_node.name,
                "Parent": parent_name,
                "Docstring": ast.get_docstring(func_node) or "",
                "Calls": ", ".join(sorted(set(filtered_calls)))
            })

    visitor = FunctionVisitor()
    visitor.visit(node)
    return structure

# 📦 프로젝트 전체 분석
def extract_project_structure(base_dir):
    ignore_dirs = {'.venv', '__pycache__', '.git'}
    records = []
    # 프로젝트 전체 심볼 수집 (Python 파일에서만)
    project_symbols = extract_project_symbols(base_dir)

    for root, dirs, files in os.walk(base_dir):
        dirs[:] = [d for d in dirs if d not in ignore_dirs]

        for file in files:
            full_path = os.path.join(root, file)
            rel_path = os.path.relpath(full_path, base_dir).replace("\\", "/")
            parts = rel_path.split("/")
            while len(parts) < 5:
                parts.append("")
            parts.append("") # File Name을 위한 추가 Level

            if file.endswith('.py'):
                local_symbols = extract_local_symbols(full_path) # Python 파일의 로컬 심볼 추출
                try:
                    functions_classes = extract_functions_and_classes(full_path, project_symbols, local_symbols)
                    for item in functions_classes:
                        records.append({
                            "Level_1": parts[0],
                            "Level_2": parts[1],
                            "Level_3": parts[2],
                            "Level_4": parts[3],
                            "Level_5": parts[4],
                            "Type": item["Type"],
                            "Name": item["Name"],
                            "Docstring": item["Docstring"],
                            "Calls": item.get("Calls", ""),
                            "File": file # 파일 이름 추가
                        })
                except Exception as e:
                    print(f"[ERROR] Failed to process Python file {rel_path}: {e}")

            elif file.endswith('.rs'): # Rust 파일 처리 로직 추가
                records.append({
                    "Level_1": parts[0],
                    "Level_2": parts[1],
                    "Level_3": parts[2],
                    "Level_4": parts[3],
                    "Level_5": parts[4],
                    "Type": "Rust File", # 타입 지정
                    "Name": file,       # 파일 이름이 곧 Name
                    "Docstring": "",    # Docstring 없음
                    "Calls": "",        # 호출 정보 없음 (AST 파싱 불가)
                    "File": file        # 파일 이름 추가
                })
            # 다른 확장자 파일을 추가하고 싶다면 여기에 elif 절 추가
            # elif file.endswith(('.txt', '.md')):
            #     records.append({
            #         "Level_1": parts[0],
            #         "Level_2": parts[1],
            #         "Level_3": parts[2],
            #         "Level_4": parts[3],
            #         "Level_5": parts[4],
            #         "Type": "Other File",
            #         "Name": file,
            #         "Docstring": "",
            #         "Calls": "",
            #         "File": file
            #     })


    return pd.DataFrame(records)

# 💾 엑셀 저장
def save_structure_to_excel(base_dir, output_file='project_structure.xlsx'):
    df = extract_project_structure(base_dir)
    # Excel 저장 전에 Level_1~5 컬럼이 없을 수도 있으므로 순서 조정
    # 'File' 컬럼도 추가했으므로 순서에 맞게 조정
    col_order = ["Level_1", "Level_2", "Level_3", "Level_4", "Level_5", "File", "Type", "Name", "Docstring", "Calls"]
    # 실제 데이터프레임에 있는 컬럼만 필터링하여 순서 적용
    df = df[df.columns.intersection(col_order)]
    # 만약 'File' 컬럼이 Level_5 다음으로 오도록 조정한다면, parts 정의도 확인
    # 현재 parts 정의에 "File"이 마지막에 추가될 수 있도록 되어 있음.

    df.to_excel(output_file, index=False)
    format_excel_docstring_and_levels(output_file)
    print(f"✅[SUCCESS] Excel saved to: {output_file}")

# Excel 포맷팅 함수 (이전과 동일)
def format_excel_docstring_and_levels(excel_path):
    wb = load_workbook(excel_path)
    ws = wb.active

    headers = {cell.value: cell.column for cell in ws[1]}
    doc_col = headers.get("Docstring")
    name_col = headers.get("Name")
    type_col = headers.get("Type")
    level_cols = [headers.get(f"Level_{i}") for i in range(1, 6)]
    file_col = headers.get("File") # 'File' 컬럼 추가

    # 병합 처리: Level_1 ~ Level_5 및 File
    merge_target_cols = level_cols + [file_col]
    for col_idx in merge_target_cols:
        if not col_idx:
            continue
        start = 2
        # Check if cell exists before accessing .value
        prev_val = ws.cell(row=2, column=col_idx).value if ws.max_row >= 2 else None
        for row in range(3, ws.max_row + 2):
            curr_val = ws.cell(row=row, column=col_idx).value if row <= ws.max_row else None
            if curr_val != prev_val:
                if row - start > 1:
                    ws.merge_cells(start_row=start, start_column=col_idx, end_row=row-1, end_column=col_idx)
                    ws.cell(row=start, column=col_idx).alignment = Alignment(
                        wrap_text=True, vertical="center", horizontal="center"
                    )
                start = row
                prev_val = curr_val

    # 모든 셀 줄바꿈 (Docstring 포함)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='center')

    # 열 너비 자동 조정 (내용 최대길이 기준)
    for col_idx, col_cells in enumerate(ws.columns, start=1):
        max_length = 0
        # 각 셀에 대해 예외 처리 추가
        for cell in col_cells:
            try:
                if cell.value:
                    # 셀 값의 문자열 길이로 처리
                    cell_len = len(str(cell.value))
                    # 줄바꿈 문자가 있는 경우, 가장 긴 줄의 길이를 고려 (간단하게 줄바꿈 제거 후 길이 계산)
                    if "\n" in str(cell.value):
                        cell_len = max(len(line) for line in str(cell.value).split('\n'))
                    max_length = max(max_length, cell_len)
            except Exception: # TypeError 등이 발생할 수 있으므로 광범위하게 처리
                pass
        adjusted_width = max_length + 2 # 여유 공간 추가
        if adjusted_width > 100: # 너무 넓어지는 것을 방지 (옵션)
            adjusted_width = 100
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    # 마지막에 Docstring 열 너비 조정: Name의 1.5배 (또는 적절한 고정 값)
    if doc_col and name_col:
        name_width = ws.column_dimensions[get_column_letter(name_col)].width or 20
        # Docstring은 내용이 길기 때문에 Name의 1.5배도 작을 수 있습니다.
        # 필요에 따라 고정 값 (예: 50)을 주거나 더 크게 조정 가능.
        ws.column_dimensions[get_column_letter(doc_col)].width = name_width * 2 # 1.5배에서 2배로 늘림

    # Level_1~5 + File + Type 열만 중앙 정렬 (wrap_text 포함)
    target_cols = [headers.get(f'Level_{i}') for i in range(1, 6)] + [headers.get("File"), headers.get("Type")]
    target_cols = [c for c in target_cols if c]  # None 제거

    for r in range(2, ws.max_row + 1):
        for c in target_cols:
            cell = ws.cell(row=r, column=c)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 저장
    enhanced_path = excel_path.replace(".xlsx", ".xlsx") # 동일 파일 덮어쓰기
    wb.save(enhanced_path)
    wb.close()
    # print(f"📁 Enhanced Excel saved to: {enhanced_path}")

# ▶️ 실행
if __name__ == '__main__':
    project_root = os.path.abspath('.')  # 현재 스크립트가 있는 디렉토리를 프로젝트 루트로 설정
    save_structure_to_excel(project_root)